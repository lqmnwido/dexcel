#!/usr/bin/env python3
"""
Dexcel — Database Schema Description Exporter
"""

import os
import platform
import subprocess
import logging
import traceback
import time
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from textual import work
from textual.app import App, ComposeResult
from textual.containers import Horizontal, Vertical, VerticalScroll
from textual.screen import Screen
from textual.widgets import Button, Input, Label, RadioButton, RadioSet, RichLog, Static


# ── Logging ──────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_DIR = os.path.join(SCRIPT_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)

_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
LOG_FILE = os.path.join(LOG_DIR, f"dexcel_{_timestamp}.log")

logger = logging.getLogger("dexcel")
logger.setLevel(logging.DEBUG)
_fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
_fh.setLevel(logging.DEBUG)
_fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
logger.addHandler(_fh)
_ch = logging.StreamHandler()
_ch.setLevel(logging.WARNING)
_ch.setFormatter(logging.Formatter("%(message)s"))
logger.addHandler(_ch)


# ── Constants ────────────────────────────────────────────────────────

DB_TYPES_ORDERED = [
    ("mysql", "MySQL / MariaDB"),
    ("postgresql", "PostgreSQL"),
    ("sqlite", "SQLite"),
    ("mssql", "Microsoft SQL Server"),
    ("oracle", "Oracle"),
]

DEFAULT_PORTS = {
    "mysql": 3306,
    "postgresql": 5432,
    "mssql": 1433,
    "oracle": 1521,
}

def open_file_location(filepath):
    filepath = os.path.abspath(filepath)
    try:
        if platform.system() == "Darwin":
            subprocess.Popen(["open", "-R", filepath])
        elif platform.system() == "Windows":
            subprocess.Popen(["explorer", "/select,", os.path.normpath(filepath)])
        else:
            subprocess.Popen(["xdg-open", os.path.dirname(filepath)])
    except Exception as exc:
        logger.error("Could not open saved file location: %s", exc)


def build_connection(driver, params):
    """Open one connection using the original, known-fast driver path."""

    if driver == "sqlite":
        path = params["path"]

        if not os.path.isfile(path):
            raise FileNotFoundError(f"SQLite file not found: {path}")

        if on_phase:
            on_phase("Loading database driver...")
        import sqlite3
        conn = sqlite3.connect(path)
        db_name = os.path.splitext(os.path.basename(path))[0]
        return conn, db_name, driver

    host = params["host"]
    port = params["port"]
    user = params["user"]
    password = params["password"]
    database = params["database"]

    logger.info(
        "Connecting: driver=%s host=%s port=%s user=%s db=%s password=%s",
        driver,
        host,
        port,
        user,
        database,
        "*" * len(password),
    )

    if driver == "mysql":
        if on_phase:
            on_phase("Loading database driver...")
        t0 = time.perf_counter()
        import pymysql
        logger.info("Driver import for mysql took %.2fs", time.perf_counter() - t0)

        conn = pymysql.connect(
            host=host,
            port=int(port),
            user=user,
            password=password,
            database=database,
            charset="utf8mb4",
            connect_timeout=10,
        )

    elif driver == "postgresql":
        if on_phase:
            on_phase("Loading database driver...")
        t0 = time.perf_counter()
        import psycopg2
        logger.info(
            "Driver import for postgresql took %.2fs",
            time.perf_counter() - t0,
        )

        conn = psycopg2.connect(
            host=host,
            port=port,
            user=user,
            password=password,
            dbname=database,
            connect_timeout=10,
        )

    elif driver == "mssql":
        if on_phase:
            on_phase("Loading database driver...")
        t0 = time.perf_counter()
        try:
            import pyodbc
        except ImportError as exc:
            raise ImportError(
                "SQL Server driver unavailable. Install Microsoft ODBC Driver 17."
            ) from exc
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={host},{port};"
            f"DATABASE={database};"
            f"UID={user};"
            f"PWD={password}"
        )
        conn = pyodbc.connect(conn_str, timeout=10)

        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={host},{port};DATABASE={database};"
            f"UID={user};PWD={password}"
        )
        if on_phase:
            on_phase("Opening database connection...")
        t0 = time.perf_counter()
        conn = pyodbc.connect(conn_str, timeout=DB_CONNECT_TIMEOUT)
        logger.info("MSSQL connect/auth took %.2fs", time.perf_counter() - t0)
    elif driver == "oracle":
        if on_phase:
            on_phase("Loading database driver...")
        t0 = time.perf_counter()
        import oracledb
        logger.info("Driver import for oracle took %.2fs", time.perf_counter() - t0)

        dsn = oracledb.makedsn(host, port, service_name=database)
        conn = oracledb.connect(user=user, password=password, dsn=dsn)

        if on_phase:
            on_phase("Opening database connection...")
        t0 = time.perf_counter()
        conn = oracledb.connect(user=user, password=password, dsn=dsn)
        logger.info("Oracle connect/auth took %.2fs", time.perf_counter() - t0)
    else:
        raise ValueError(f"Unsupported driver: {driver}")

    return conn, database, driver


def list_tables(conn, driver: str):
    cursor = conn.cursor()
    if driver == "mysql":
        cursor.execute("SHOW FULL TABLES WHERE Table_type = 'BASE TABLE'")
        return [row[0] for row in cursor.fetchall()]
    if driver == "postgresql":
        cursor.execute(
            "SELECT table_name FROM information_schema.tables "
            "WHERE table_schema = 'public' AND table_type = 'BASE TABLE' "
            "ORDER BY table_name"
        )
        return [row[0] for row in cursor.fetchall()]
    if driver == "sqlite":
        cursor.execute(
            "SELECT name FROM sqlite_master "
            "WHERE type = 'table' AND name NOT LIKE 'sqlite_%' "
            "ORDER BY name"
        )
        return [row[0] for row in cursor.fetchall()]
    if driver == "mssql":
        cursor.execute(
            "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES "
            "WHERE TABLE_TYPE = 'BASE TABLE' ORDER BY TABLE_NAME"
        )
        return [row[0] for row in cursor.fetchall()]
    if driver == "oracle":
        cursor.execute("SELECT table_name FROM user_tables ORDER BY table_name")
        return [row[0] for row in cursor.fetchall()]
    raise ValueError(f"Unsupported driver: {driver}")


def get_table_description(conn, driver: str, table: str):
    cursor = conn.cursor()

    if driver == "mysql":
        cursor.execute(f"DESCRIBE `{table}`")
        return [
            {"Field": r[0], "Type": r[1], "Null": r[2], "Key": r[3],
             "Default": "NULL" if r[4] is None else r[4], "Extra": r[5]}
            for r in cursor.fetchall()
        ]

    if driver == "postgresql":
        cursor.execute("""
            SELECT c.column_name, c.data_type, c.is_nullable,
                   CASE WHEN pk.column_name IS NOT NULL THEN 'PRI' ELSE '' END,
                   c.column_default,
                   CASE WHEN c.column_default LIKE 'nextval%%' THEN 'auto_increment' ELSE '' END
            FROM information_schema.columns c
            LEFT JOIN (
                SELECT ku.table_name, ku.column_name
                FROM information_schema.table_constraints tc
                JOIN information_schema.key_column_usage ku
                    ON tc.constraint_name = ku.constraint_name
                WHERE tc.constraint_type = 'PRIMARY KEY' AND ku.table_schema = 'public'
            ) pk ON c.table_name = pk.table_name AND c.column_name = pk.column_name
            WHERE c.table_schema = 'public' AND c.table_name = %s
            ORDER BY c.ordinal_position
        """, (table,))
        return [
            {"Field": r[0], "Type": r[1], "Null": "YES" if r[2] == "YES" else "NO",
             "Key": r[3] or "", "Default": "NULL" if r[4] is None else r[4],
             "Extra": r[5] or ""}
            for r in cursor.fetchall()
        ]

    if driver == "sqlite":
        cursor.execute(f'PRAGMA table_info("{table}")')
        return [
            {"Field": r[1], "Type": r[2], "Null": "NO" if r[3] else "YES",
             "Key": "PRI" if r[5] else "", "Default": "NULL" if r[4] is None else r[4],
             "Extra": ""}
            for r in cursor.fetchall()
        ]

    if driver == "mssql":
        cursor.execute("""
            SELECT c.COLUMN_NAME,
                   c.DATA_TYPE + CASE WHEN c.CHARACTER_MAXIMUM_LENGTH IS NOT NULL
                       THEN '(' + CAST(c.CHARACTER_MAXIMUM_LENGTH AS VARCHAR) + ')' ELSE '' END,
                   c.IS_NULLABLE,
                   CASE WHEN pk.COLUMN_NAME IS NOT NULL THEN 'PRI' ELSE '' END,
                   c.COLUMN_DEFAULT,
                   CASE WHEN COLUMNPROPERTY(OBJECT_ID(c.TABLE_SCHEMA+'.'+c.TABLE_NAME),
                       c.COLUMN_NAME, 'IsIdentity') = 1 THEN 'auto_increment' ELSE '' END
            FROM INFORMATION_SCHEMA.COLUMNS c
            LEFT JOIN (
                SELECT ku.TABLE_NAME, ku.COLUMN_NAME
                FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS tc
                JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE ku
                    ON tc.CONSTRAINT_NAME = ku.CONSTRAINT_NAME
                WHERE tc.CONSTRAINT_TYPE = 'PRIMARY KEY'
            ) pk ON c.TABLE_NAME = pk.TABLE_NAME AND c.COLUMN_NAME = pk.COLUMN_NAME
            WHERE c.TABLE_NAME = ?
            ORDER BY c.ORDINAL_POSITION
        """, table)
        return [
            {"Field": r[0], "Type": r[1], "Null": "YES" if r[2] == "YES" else "NO",
             "Key": r[3] or "", "Default": "NULL" if r[4] is None else r[4],
             "Extra": r[5] or ""}
            for r in cursor.fetchall()
        ]

    if driver == "oracle":
        cursor.execute("""
            SELECT c.column_name,
                   c.data_type || CASE
                       WHEN c.data_type IN ('VARCHAR2','CHAR','NVARCHAR2','NCHAR')
                           THEN '(' || c.data_length || ')'
                       WHEN c.data_type = 'NUMBER' AND c.data_precision IS NOT NULL
                           THEN '(' || c.data_precision || ',' || c.data_scale || ')'
                       ELSE '' END,
                   c.nullable,
                   CASE WHEN pk.column_name IS NOT NULL THEN 'PRI' ELSE '' END,
                   c.data_default, ''
            FROM user_tab_columns c
            LEFT JOIN (
                SELECT cols.table_name, cols.column_name
                FROM user_constraints cons
                JOIN user_cons_columns cols
                    ON cons.constraint_name = cols.constraint_name
                WHERE cons.constraint_type = 'P'
            ) pk ON c.table_name = pk.table_name AND c.column_name = pk.column_name
            WHERE c.table_name = :table_name
            ORDER BY c.column_id
        """, table_name=table.upper())
        return [
            {"Field": r[0], "Type": r[1],
             "Null": "YES" if r[2] == "Y" else "NO",
             "Key": r[3] or "",
             "Default": "NULL" if r[4] is None else str(r[4]).strip(),
             "Extra": r[5] or ""}
            for r in cursor.fetchall()
        ]

    raise ValueError(f"Unsupported driver: {driver}")


def make_box_line(widths):
    return "+" + "+".join("-" * (width + 2) for width in widths) + "+"


def make_box_row(values, widths):
    cells = []

    for value, width in zip(values, widths):
        text = "" if value is None else str(value)
        cells.append(" " + text.ljust(width) + " ")

    return "|" + "|".join(cells) + "|"


def render_table_box(headers, rows):
    normalized_rows = []

    for row in rows:
        normalized_rows.append([row.get(header, "") for header in headers])

    widths = []

    for index, header in enumerate(headers):
        max_width = len(header)

        for row in normalized_rows:
            max_width = max(max_width, len(str(row[index])))

        widths.append(max_width)

    output = []
    line = make_box_line(widths)

    output.append(line)
    output.append(make_box_row(headers, widths))
    output.append(line)

    for row in normalized_rows:
        output.append(make_box_row(row, widths))

    output.append(line)

    return "\n".join(output)


def render_table_title(table_name):
    title = f"TABLE: {table_name}"
    width = len(title)

    return "\n".join(
        [
            "+" + "-" * (width + 2) + "+",
            "| " + title + " |",
            "+" + "-" * (width + 2) + "+",
        ]
    )


def export_schema_descriptions(
    conn,
    driver,
    tables,
    output_file,
    on_progress=None,
    on_saving=None,
):
    headers = ["Field", "Type", "Null", "Key", "Default", "Extra"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Table Descriptions"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(color="FFFFFF", bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_al = Alignment(horizontal="center", vertical="center")
    left_al = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row_num = 1
    for table in tables:
        if on_progress:
            on_progress(table)
        logger.info("Describing table: %s", table)

        try:
            t0 = time.perf_counter()
            rows = get_table_description(conn, driver, table)
            logger.info(
                "Description for table %s took %.2fs",
                table,
                time.perf_counter() - t0,
            )
        except Exception as e:
            logger.exception("Could not describe table %s", table)
            ws.merge_cells(
                start_row=row_num, start_column=1,
                end_row=row_num, end_column=len(HEADERS),
            )
            cell = ws.cell(row=row_num, column=1, value=f"TABLE: {table}")
            cell.fill = title_fill
            cell.font = title_font
            cell.alignment = left_al
            row_num += 1
            ws.cell(row=row_num, column=1, value=f"ERROR: {e}")
            row_num += 2
            continue

        ws.merge_cells(
            start_row=row_num, start_column=1,
            end_row=row_num, end_column=len(HEADERS),
        )
        tc = ws.cell(row=row_num, column=1, value=f"TABLE: {table}")
        tc.fill = title_fill
        tc.font = title_font
        tc.alignment = left_al
        tc.border = border
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=row_num, column=c).border = border
        row_num += 1

        for cn, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=row_num, column=cn, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_al
        row_num += 1

        for row in rows:
            for cn, h in enumerate(HEADERS, 1):
                v = row.get(h, "") or "NULL"
                cell = ws.cell(row=row_num, column=cn, value=str(v))
                cell.border = border
                cell.alignment = left_al
            row_num += 1
        row_num += 1

    for cc in ws.columns:
        ml = max((len(str(c.value or "")) for c in cc), default=0)
        ws.column_dimensions[get_column_letter(cc[0].column)].width = min(ml + 2, 60)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25

    if on_saving:
        on_saving()
    wb.save(output_file)


# ── Screens ─────────────────────────────────────────────────────────

class MainScreen(Screen):
    def compose(self) -> ComposeResult:
        yield Static("Dexcel — Database Schema Exporter", id="titlebar")
        with Vertical(id="page"):
            yield Label("Database Reader", id="screen-title")
            yield Label("Select a database type to describe.")
            with RadioSet(id="db-type"):
                for key, name in DB_TYPES_ORDERED:
                    yield RadioButton(f"{name}", id=key)
            with Horizontal():
                yield Button("Next", variant="primary", id="btn-next", disabled=True)

    def on_radio_set_changed(self, event: RadioSet.Changed) -> None:
        self.app.driver = event.pressed.id
        self.query_one("#btn-next", Button).disabled = False

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "btn-next":
            d = self.app.driver
            if d == "sqlite":
                self.app.push_screen(SQLiteScreen())
            else:
                self.app.push_screen(NetScreen())


class SQLiteScreen(Screen):
    def compose(self) -> ComposeResult:
        yield Static("Dexcel — Database Schema Exporter", id="titlebar")
        with Vertical(id="page"):
            yield Label("SQLite Database File", id="screen-title")
            yield Label("Path")
            yield Input(placeholder="Path to .db or .sqlite file", id="db-path")
            yield Label("", id="sqlite-error")
            with Horizontal():
                yield Button("Cancel", id="back")
                yield Button("Describe", variant="primary", id="connect")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "back":
            self.app.pop_screen()
        elif event.button.id == "connect":
            path = self.query_one("#db-path", Input).value.strip()
            if not path:
                self.query_one("#sqlite-error", Label).update("Enter a file path.")
                return
            if not os.path.isfile(path):
                self.query_one("#sqlite-error", Label).update(f"File not found: {path}")
                return
            self.app.db_params = {"path": path}
            self.app.push_screen(ExportScreen())


class NetScreen(Screen):
    def compose(self) -> ComposeResult:
        yield Static("Dexcel — Database Schema Exporter", id="titlebar")
        with Vertical(id="page"):
            yield Label("Database Connection", id="screen-title")
            yield Label("Host")
            yield Input(placeholder="Host", id="host")
            yield Label("Port")
            yield Input(placeholder="Port", id="port")
            yield Label("Username")
            yield Input(placeholder="Username", id="username")
            yield Label("Password")
            yield Input(placeholder="Password", id="password", password=True)
            yield Label("Dataname")
            yield Input(placeholder="Dataname / Database name", id="database")
            yield Label("", id="net-error")
            with Horizontal():
                yield Button("Cancel", id="back")
                yield Button("Approve Describe", variant="primary", id="connect")

    def on_mount(self) -> None:
        driver = self.app.driver
        self.query_one("#host", Input).value = "127.0.0.1"
        if driver == "mysql":
            self.query_one("#username", Input).value = "root"
        port = DEFAULT_PORTS.get(driver)
        if port:
            self.query_one("#port", Input).value = str(port)

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "back":
            self.app.pop_screen()
        elif event.button.id == "connect":
            fields = {
                "host": self.query_one("#host", Input).value.strip(),
                "port": self.query_one("#port", Input).value.strip(),
                "user": self.query_one("#username", Input).value.strip(),
                "password": self.query_one("#password", Input).value,
                "database": self.query_one("#database", Input).value.strip(),
            }
            missing = [k for k, v in fields.items() if not v]
            if missing:
                self.query_one("#net-error", Label).update(
                    f"Missing required: {', '.join(missing)}"
                )
                return
            self.app.db_params = fields
            self.app.push_screen(ExportScreen())


class ExportScreen(Screen):
    def compose(self) -> ComposeResult:
        yield Static("Dexcel — Database Schema Exporter", id="titlebar")
        with Vertical(id="page"):
            yield Label("Export Progress")
            yield Label("Preparing export...", id="export-status")
            yield RichLog(id="log", highlight=True, markup=True)
            yield Label("Saved Excel file")
            yield Button("Waiting for export...", id="file-link", disabled=True)
            with Horizontal():
                yield Button("Cancel", id="cancel")
                yield Button("Exit", variant="primary", id="exit", disabled=True)

    def on_mount(self) -> None:
        self.query_one("#export-status", Label).update("Preparing export...")
        self.call_after_refresh(self.run_export)

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "cancel":
            self.app.pop_screen()
        elif event.button.id == "exit":
            self.app.exit()
        elif event.button.id == "file-link":
            if self.app.output_file:
                open_file_location(self.app.output_file)

    @work(thread=True, exit_on_error=False)
    def run_export(self) -> None:
        def _log(msg: str) -> None:
            self.call_from_thread(self._append_log, msg)

        def _status(msg: str) -> None:
            self.call_from_thread(self._update_status, msg)

        try:
            _log(f"Diagnostic log: {LOG_FILE}")

            def connection_phase(msg: str) -> None:
                _status(msg)
                _log(msg)

            conn, db_name = build_connection(
                self.app.driver,
                on_phase=connection_phase,
                **self.app.db_params,
            )
            self.app.conn = conn
            self.app.db_name = db_name
            _status("Connection established.")
            _log("Connected successfully.")

            _status("Reading table list...")
            _log("Reading table list...")
            t0 = time.perf_counter()
            tables = list_tables(conn, self.app.driver)
            logger.info("list_tables took %.2fs", time.perf_counter() - t0)
            _log(f"Found {len(tables)} table(s).")

            if not tables:
                _log("No tables found in this database.")
                self.call_from_thread(self._finish, False)
                return

            output_file = f"{db_name}_table_descriptions.xlsx"
            output_path = os.path.abspath(output_file)
            self.app.output_file = output_path

            if os.path.exists(output_path):
                _log("File exists -- will overwrite")

            _status("Reading table descriptions...")
            _log(f"Output: {output_path}")

            def on_progress(msg: str) -> None:
                self.call_from_thread(self._append_log, f"  {msg}")

            export_started = time.perf_counter()
            export_to_excel(
                conn,
                self.app.driver,
                tables,
                output_path,
                on_progress,
                on_phase=_status,
            )
            logger.info(
                "Total schema export took %.2fs",
                time.perf_counter() - export_started,
            )

            _status("Export complete.")
            _log("Export complete!")
            self.call_from_thread(self._finish, True, output_path)

        except Exception as e:
            _status("Export failed.")
            _log(f"Error: {e}")
            logger.error("Export failed: %s", traceback.format_exc())
            self.call_from_thread(self._finish, False, error=str(e))
        finally:
            try:
                if hasattr(self.app, "conn") and self.app.conn:
                    self.app.conn.close()
            except Exception:
                pass

    def _append_log(self, message: str) -> None:
        self.query_one("#log", RichLog).write(message)

    def _update_status(self, message: str) -> None:
        self.query_one("#export-status", Label).update(message)

    def _finish(self, success: bool, file_path: Optional[str] = None,
                error: Optional[str] = None) -> None:
        self.query_one("#cancel", Button).disabled = True

        if success and file_path:
            self.query_one("#exit", Button).disabled = False
            file_link = self.query_one("#file-link", Button)
            file_link.disabled = False
            file_link.label = file_path
        elif error:
            self.query_one("#exit", Button).disabled = False
            file_link = self.query_one("#file-link", Button)
            file_link.label = "Export failed. Check the log above."
            file_link.disabled = True
        else:
            self.query_one("#exit", Button).disabled = False
            file_link = self.query_one("#file-link", Button)
            file_link.label = "Waiting for export..."
            file_link.disabled = True


# ── App ──────────────────────────────────────────────────────────────

class DexcelApp(App):
    """Dexcel — Database Schema Exporter"""

    CSS = """
    Screen {
        background: #191d28;
    }

    #titlebar {
        background: #7c5cff;
        color: #ffffff;
        text-style: bold;
        height: 1;
        content-align: center middle;
        width: 100%;
    }

    #page {
        align: left top;
        width: 100%;
        height: 1fr;
        margin: 0;
        padding: 1 2;
        border: round #394056;
        background: #222836;
        overflow: auto;
    }

    Input {
        margin: 0 0 1 0;
        width: 100%;
    }

    #screen-title {
        text-style: bold;
        color: #f2f4ff;
        margin: 0 0 1 0;
    }

    RichLog {
        height: 1fr;
        min-height: 10;
        margin: 0 0 1 0;
        border: round #394056;
        background: #111521;
    }

    Horizontal {
        align: center middle;
        width: 100%;
        height: auto;
        margin: 1 0 0 0;
    }

    Horizontal Button {
        margin: 0 1;
        min-width: 12;
    }

    #file-link {
        width: 100%;
    }

    #net-error,
    #sqlite-error {
        color: #ff8a8a;
        margin: 0 0 1 0;
    }
    """

    def __init__(self):
        super().__init__()
        self.driver = None
        self.db_params = {}
        self.conn = None
        self.db_name = None
        self.output_file = None

    def on_mount(self) -> None:
        self.push_screen(MainScreen())


class ExportCancelled(Exception):
    pass


class DatabaseScreen(Screen):
    def compose(self) -> ComposeResult:
        yield Static("DEXCEL  /  DATABASE READER", classes="topbar")
        with VerticalScroll(classes="page"):
            yield Label("Choose a database", classes="heading")
            yield Label(
                "Select the engine whose schema you want to describe.",
                classes="muted",
            )
            with RadioSet(id="database-types"):
                for info in DB_TYPES.values():
                    yield RadioButton(info["name"], id=info["driver"])
            with Horizontal(classes="actions"):
                yield Button("Exit", id="exit")
                yield Button("Continue", id="continue", variant="primary", disabled=True)

    def on_radio_set_changed(self, event: RadioSet.Changed) -> None:
        if event.pressed is None:
            return
        self.app.driver = event.pressed.id
        self.query_one("#continue", Button).disabled = False

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "exit":
            self.app.exit()
        elif event.button.id == "continue":
            self.app.push_screen(ConnectionScreen(self.app.driver))


class ConnectionScreen(Screen):
    def __init__(self, driver):
        super().__init__()
        self.driver = driver

    def compose(self) -> ComposeResult:
        yield Static("DEXCEL  /  CONNECTION", classes="topbar")
        with VerticalScroll(classes="page"):
            yield Label("Database connection", classes="heading")
            yield Label(
                next(info["name"] for info in DB_TYPES.values()
                     if info["driver"] == self.driver),
                classes="engine-name",
            )
            if self.driver == "sqlite":
                yield Label("Database file")
                yield Input(
                    placeholder="/path/to/database.sqlite",
                    id="database-path",
                )
            else:
                yield Label("Host")
                yield Input(value="127.0.0.1", placeholder="Database host", id="host")
                yield Label("Port")
                yield Input(
                    value=str(DEFAULT_PORTS[self.driver]),
                    placeholder="Port",
                    id="port",
                )
                yield Label("Username")
                yield Input(
                    value="root" if self.driver == "mysql" else "",
                    placeholder="Username",
                    id="username",
                )
                yield Label("Password")
                yield Input(
                    placeholder="Password (may be empty)",
                    password=True,
                    id="password",
                )
                yield Label("Dataname")
                yield Input(
                    placeholder="Database name or service name",
                    id="database",
                )
            yield Label("", id="form-error")
            with Horizontal(classes="actions"):
                yield Button("Cancel", id="cancel")
                yield Button(
                    "Approve Describe",
                    id="describe",
                    variant="primary",
                )

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "cancel":
            self.app.pop_screen()
            return
        if event.button.id != "describe":
            return

        error = self.query_one("#form-error", Label)
        if self.driver == "sqlite":
            path = os.path.expanduser(
                self.query_one("#database-path", Input).value.strip()
            )
            if not path:
                error.update("Database file is required.")
                return
            if not os.path.isfile(path):
                error.update("Database file does not exist.")
                return
            params = {"path": path}
        else:
            params = {
                "host": self.query_one("#host", Input).value.strip(),
                "port": self.query_one("#port", Input).value.strip(),
                "user": self.query_one("#username", Input).value.strip(),
                "password": self.query_one("#password", Input).value,
                "database": self.query_one("#database", Input).value.strip(),
            }
            missing = [
                name for name in ("host", "port", "user", "database")
                if not params[name]
            ]
            if missing:
                error.update(f"Required: {', '.join(missing)}")
                return
            try:
                int(params["port"])
            except ValueError:
                error.update("Port must be a number.")
                return

        self.app.db_params = params
        self.app.push_screen(ExportScreen())


class ExportScreen(Screen):
    def __init__(self):
        super().__init__()
        self.cancel_requested = False
        self.running = True

    def compose(self) -> ComposeResult:
        yield Static("DEXCEL  /  DATABASE DESCRIBE", classes="topbar")
        with Vertical(classes="export-page"):
            yield Label("Preparing export", classes="heading", id="status")
            yield Label("The active operation is shown below.", classes="muted")
            yield RichLog(id="progress-log", markup=False, wrap=True)
            yield Label("Saved Excel file", classes="section-label")
            yield Button(
                "Waiting for export...",
                id="saved-file",
                disabled=True,
            )
            with Horizontal(classes="actions"):
                yield Button("Cancel", id="cancel")
                yield Button("Exit", id="exit", variant="primary", disabled=True)

    def on_mount(self) -> None:
        self.run_export()

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "saved-file" and self.app.output_file:
            open_file_location(self.app.output_file)
        elif event.button.id == "exit":
            self.app.exit()
        elif event.button.id == "cancel":
            if self.running:
                self.cancel_requested = True
                event.button.disabled = True
                self.query_one("#status", Label).update("Cancelling safely...")
            else:
                self.app.pop_screen()

    @work(thread=True, exit_on_error=False)
    def run_export(self) -> None:
        conn = None
        try:
            self._thread_status("Loading driver and opening connection...")
            conn, db_name, driver = build_connection(
                self.app.driver,
                self.app.db_params,
            )
            self._thread_log("Connected successfully.")

            if self.cancel_requested:
                raise ExportCancelled()

            self._thread_status("Reading table list...")
            tables = list_tables(conn, driver)
            self._thread_log(f"Found {len(tables)} table(s).")
            if not tables:
                raise RuntimeError("No tables were found in this database.")

            output_path = os.path.abspath(
                f"{db_name}_table_descriptions.xlsx"
            )
            self.app.output_file = output_path
            if os.path.exists(output_path):
                self._thread_log("Existing Excel file will be replaced.")

            self._thread_status("Reading table descriptions...")

            def on_progress(table):
                if self.cancel_requested:
                    raise ExportCancelled()
                self._thread_log(f"Describing: {table}")

            export_schema_descriptions(
                conn,
                driver,
                tables,
                output_path,
                on_progress=on_progress,
                on_saving=lambda: self._thread_status("Saving Excel file..."),
            )

            self._thread_status("Export complete")
            self._thread_log(f"Saved: {output_path}")
            logger.info("=== Dexcel schema export session finished successfully ===")
            self.app.call_from_thread(self._finish, True, output_path, None)
        except ExportCancelled:
            self._thread_status("Export cancelled")
            self._thread_log("The operation was cancelled.")
            self.app.call_from_thread(self._finish, False, None, None)
        except Exception as exc:
            logger.error("Export failed:\n%s", traceback.format_exc())
            self._thread_status("Export failed")
            self._thread_log(f"Error: {exc}")
            self._thread_log(f"Diagnostic log: {LOG_FILE}")
            self.app.call_from_thread(self._finish, False, None, str(exc))
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    logger.warning("Could not close database connection.")

    def _thread_status(self, message):
        self.app.call_from_thread(self._set_status, message)

    def _thread_log(self, message):
        self.app.call_from_thread(self._write_log, message)

    def _set_status(self, message):
        self.query_one("#status", Label).update(message)

    def _write_log(self, message):
        self.query_one("#progress-log", RichLog).write(message)

    def _finish(
        self,
        success,
        output_path: Optional[str],
        error: Optional[str],
    ) -> None:
        self.running = False
        cancel = self.query_one("#cancel", Button)
        cancel.disabled = False
        cancel.label = "Back"
        self.query_one("#exit", Button).disabled = False

        saved_file = self.query_one("#saved-file", Button)
        if success and output_path:
            saved_file.label = output_path
            saved_file.disabled = False
        elif error:
            saved_file.label = "Export failed. Review the progress log."
        else:
            saved_file.label = "No Excel file was created."


class DexcelApp(App):
    TITLE = "Dexcel"
    ENABLE_COMMAND_PALETTE = False

    CSS = """
    Screen {
        background: #10161d;
        color: #e7edf2;
    }

    .topbar {
        width: 100%;
        height: 3;
        padding: 1 2 0 2;
        background: #0d5661;
        color: #f3fbf8;
        text-style: bold;
    }

    .page, .export-page {
        width: 100%;
        height: 1fr;
        padding: 1 3;
        background: #17212b;
    }

    .heading {
        width: 100%;
        height: auto;
        color: #f4b860;
        text-style: bold;
        margin-bottom: 1;
    }

    .muted {
        width: 100%;
        height: auto;
        color: #91a3b0;
        margin-bottom: 1;
    }

    .engine-name {
        color: #55c2b2;
        text-style: bold;
        margin-bottom: 1;
    }

    RadioSet {
        width: 100%;
        height: auto;
        margin-bottom: 1;
        border: round #304552;
        background: #111a22;
    }

    Input {
        width: 100%;
        margin-bottom: 1;
        border: tall #304552;
        background: #0e171e;
    }

    Input:focus {
        border: tall #55c2b2;
    }

    #form-error {
        width: 100%;
        height: auto;
        color: #ff7b72;
    }

    .actions {
        width: 100%;
        height: 3;
        margin-top: 1;
        align: center middle;
    }

    .actions Button {
        width: 1fr;
        min-width: 10;
        margin: 0 1;
    }

    #progress-log {
        width: 100%;
        height: 1fr;
        min-height: 5;
        margin-bottom: 1;
        border: round #304552;
        background: #0e171e;
        scrollbar-color: #55c2b2;
    }

    .section-label {
        width: 100%;
        height: 1;
        color: #91a3b0;
    }

    #saved-file {
        width: 100%;
        height: 3;
        margin-top: 1;
    }
    """

    def __init__(self):
        super().__init__()
        self.driver = None
        self.db_params = {}
        self.output_file = None

    def on_mount(self) -> None:
        logger.info("=== Dexcel schema export session started ===")
        self.push_screen(DatabaseScreen())


def main():
    DexcelApp().run()


if __name__ == "__main__":
    main()
